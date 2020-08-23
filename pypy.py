from openpyxl import Workbook

wb= Workbook()

sheet=wb.active
sheet.title ='BOB'

sheet['A1']='Hello'
sheet['B1']='BoB'
wb.create_sheet('Sheet2')
wb.create_sheet('Sheet3')
wb.save('bob.xlsx')
